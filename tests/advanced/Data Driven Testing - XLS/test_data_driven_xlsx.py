from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

#to interact with the excel sheet
from openpyxl import load_workbook
from selenium.webdriver.common.by import By


def test_data_driven_xlsx(driver):
    # loading the Excel Sheet
    workbook = load_workbook('advanced/Data Driven Testing/testdata.xlsx')

    # Selecting active sheet
    sheet = workbook.active

    # iterating the sheet
    # In basic English in case I forget: For every row in the iteration of this sheet, (minimum row 2, maximum infinite but only the ones that have values):
    # the username is the row 0 and the password is the row 1
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        wait = WebDriverWait(driver, 10)
        user_name = row[0]
        password = row[1]

        driver.get('https://www.saucedemo.com/')
        wait.until(EC.presence_of_element_located((By.ID, 'user-name')))

        driver.find_element(By.ID, 'user-name').send_keys(user_name)
        driver.find_element(By.ID, 'password').send_keys(password)
        driver.find_element(By.ID, 'login-button').click()
        wait.until(EC.presence_of_element_located((By.XPATH, '//button[@id="react-burger-menu-btn"]')))
        assert driver.find_element(By.CLASS_NAME, "title").text == "Products"

        driver.find_element(By.XPATH, '//button[@id="react-burger-menu-btn"]').click()
        wait.until(EC.presence_of_element_located((By.ID, 'logout_sidebar_link')))

        driver.find_element(By.ID, 'logout_sidebar_link').click()
        assert driver.current_url == 'https://www.saucedemo.com/'


